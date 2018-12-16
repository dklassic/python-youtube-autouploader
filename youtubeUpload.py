#!/usr/bin/python

import argparse
import httplib
import httplib2
import os
import sys
import io
import random
import time

import google.oauth2.credentials
import google_auth_oauthlib.flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook

# Explicitly tell the underlying HTTP transport library not to retry, since
# we are handling retry logic ourselves.
httplib2.RETRIES = 1

# Maximum number of times to retry before giving up.
MAX_RETRIES = 10

# Always retry when these exceptions are raised.
RETRIABLE_EXCEPTIONS = (httplib2.HttpLib2Error, IOError, httplib.NotConnected,
  httplib.IncompleteRead, httplib.ImproperConnectionState,
  httplib.CannotSendRequest, httplib.CannotSendHeader,
  httplib.ResponseNotReady, httplib.BadStatusLine)

# Always retry when an apiclient.errors.HttpError with one of these status
# codes is raised.
RETRIABLE_STATUS_CODES = [500, 502, 503, 504]

# The CLIENT_SECRETS_FILE variable specifies the name of a file that contains
# the OAuth 2.0 information for this application, including its client_id and
# client_secret. You can acquire an OAuth 2.0 client ID and client secret from
# the {{ Google Cloud Console }} at
# {{ https://cloud.google.com/console }}.
# Please ensure that you have enabled the YouTube Data API for your project.
# For more information about using OAuth2 to access the YouTube Data API, see:
#   https://developers.google.com/youtube/v3/guides/authentication
# For more information about the client_secrets.json file format, see:
#   https://developers.google.com/api-client-library/python/guide/aaa_client_secrets
CLIENT_SECRETS_FILE = 'client_secret.json'

# This OAuth 2.0 access scope allows an application to upload files to the
# authenticated user's YouTube channel, but doesn't allow other types of access.
SCOPES = ["https://www.googleapis.com/auth/youtube.upload", "https://www.googleapis.com/auth/youtube", "https://www.googleapis.com/auth/youtube.force-ssl"]
API_SERVICE_NAME = 'youtube'
API_VERSION = 'v3'

VALID_PRIVACY_STATUSES = ('public', 'private', 'unlisted')


# Authorize the request and store authorization credentials.
def get_authenticated_service():
  flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
  try:
    os.remove(CLIENT_SECRETS_FILE)
  except:
    pass
  credentials = flow.run_console()
  return build(API_SERVICE_NAME, API_VERSION, credentials = credentials)


def initialize_upload(youtube, options):
  # Load project name
  name = options.name
  # Define filenames
  thumbnailFile = name + ".jpg"
  videoFile = name + ".mp4"
  videoKit = name + ".xlsx"
  TCHsrt = name + ".zh_TW.srt"
  THAIsrt = name + ".th.srt"

  # Load videokit
  wb = load_workbook(videoKit)
  ws = wb[wb.sheetnames[0]]
  # Build ENG title and description
  ENGtitle =ws["B8"].value
  if ENGtitle is not None:
    if ws["B10"].value is not None:
      ENGdescription = ws["B10"].value + "\n\n" + ws["B11"].value + "\n" + ws["B12"].value
    else:
      ENGdescription = ws["B11"].value + "\n" + ws["B12"].value
  print("Video name is " + ENGtitle)

  # Build TCH title and description if exists
  TCHtitle = ws["C8"].value
  if TCHtitle is not None:
    if ws["C10"].value is not None:
      TCHdescription = ws["C10"].value + "\n\n" + ws["C11"].value + "\n" + ws["C12"].value
    else:
      TCHdescription = ws["C11"].value + "\n" + ws["C12"].value

  # Build THAI title and description if exists
  THAItitle = ws["D8"].value
  if THAItitle is not None:
    if ws["D10"].value is not None:
      THAIdescription = ws["D10"].value + "\n\n" + ws["D11"].value + "\n" + ws["D12"].value
    else:
      THAIdescription = ws["D11"].value + "\n" + ws["D12"].value

  # tags = None
  # if options.keywords:
  #   tags = options.keywords.split(',')

  body=dict(
    snippet=dict(
      title=ENGtitle,
      description=ENGdescription,
      # tags=tags,
      categoryId=options.category,
      defaultLanguage="en"
    ),
    status=dict(
      privacyStatus=options.privacyStatus
    ),
  )

  body["localizations"] = {}
  # Add TCH is exists
  if TCHtitle is not None:
    body["localizations"]['zh-Hant'] ={
      "title": TCHtitle,
      "description": TCHdescription
    }
  # Add THAI is exists
  if THAItitle is not None:
    body["localizations"]['th'] ={
      "title": THAItitle,
      "description": THAIdescription
    }
  # Call the API's videos.insert method to create and upload the video.
  insert_request = youtube.videos().insert(
    part=','.join(body.keys()),
    body=body,
    # The chunksize parameter specifies the size of each chunk of data, in
    # bytes, that will be uploaded at a time. Set a higher value for
    # reliable connections as fewer chunks lead to faster uploads. Set a lower
    # value for better recovery on less reliable connections.
    #
    # Setting 'chunksize' equal to -1 in the code below means that the entire
    # file will be uploaded in a single HTTP request. (If the upload fails,
    # it will still be retried where it left off.) This is usually a best
    # practice, but if you're using Python older than 2.6 or if you're
    # running on App Engine, you should set the chunksize to something like
    # 1024 * 1024 (1 megabyte).
    media_body=MediaFileUpload(videoFile, chunksize=-1, resumable=True)
  )

  # Perform Upload
  video_id = resumable_upload(insert_request)
  hasTCHsrt = os.path.isfile(TCHsrt)
  hasTHAIsrt = os.path.isfile(THAIsrt)
  hasThumbnail = os.path.isfile(thumbnailFile)
  shouldWait = hasTCHsrt or hasTHAIsrt or hasThumbnail
  # Perform other operations after uploaded
  if (video_id is not None) and shouldWait:
    print("Detected files to upload, waiting the video to be processed.")
    print("Video status will be checked every 30 seconds automatically.")
    print("You can press Ctrl + C anytime to simply skip the wait and manual upload later.")
    video_status = video_upload_status(youtube, video_id)
    while video_status != "processed":
      time.sleep(30)
      video_status = video_upload_status(youtube, video_id)
    print("Video processed! Continue the procedure...")
    if hasTCHsrt:
      upload_caption(youtube, video_id, "zh-Hant", TCHsrt)
    if hasTHAIsrt:
      upload_caption(youtube, video_id, "th", THAIsrt)
    if hasThumbnail:
      upload_thumbnail(youtube, video_id, thumbnailFile)
  elif (video_id is not None) and not shouldWait:
    print("No other files to upload, great job.")

def video_upload_status(youtube, video_id):
    status_dict = youtube.videos().list(part="status", id=video_id, maxResults=5).execute()
    upload_status = status_dict['items'][0]['status']['uploadStatus']
    return upload_status

def upload_thumbnail(youtube, video_id, file):
  youtube.thumbnails().set(
    videoId=video_id,
    media_body=file
  ).execute()
  print("Thumbnail uploaded!")

def upload_caption(youtube, video_id, language, file):
  with io.open(file, 'r', encoding="utf-8") as text:
    string_return = text.read()
  #making a temporary file as google seems to only acept .txt file
  filename = language + ".txt"
  os.rename(file, filename)

  insert_result = youtube.captions().insert(
    part="snippet",
    body=dict(
      snippet=dict(
        videoId=video_id,
        language=language,
        name=" ",
        isDraft=False
      )
    ),
    media_body=filename
  ).execute()

  theid = insert_result["id"]
  name = insert_result["snippet"]["name"]
  language = insert_result["snippet"]["language"]
  status = insert_result["snippet"]["status"]

  try:
    os.rename(filename, file)
  except:
    pass
  if status != "failed":
    print(language + " subtitle has uploaded.")
  else:
    print("There may be an issue for " + language + ", please upload manually.")
  #print "Uploaded caption track '%s(%s) in '%s' language, '%s' status." % (name, theid, language, status)

def delete_caption(youtube, caption_id):
  youtube.captions().delete(
    id=caption_id
  ).execute()

  print "caption track '%s' deleted succesfully" % (caption_id)


# This method implements an exponential backoff strategy to resume a
# failed upload.
def resumable_upload(request):
  response = None
  error = None
  retry = 0
  returnvalue = None
  while response is None:
    try:
      print 'Uploading file...'
      status, response = request.next_chunk()
      if response is not None:
        if 'id' in response:
          #print 'Video id "%s" was successfully uploaded.' % response['id']
          print 'Video was successfully uploaded.'
          returnvalue = response['id']
        else:
          exit('The upload failed with an unexpected response: %s' % response)
    except HttpError, e:
      if e.resp.status in RETRIABLE_STATUS_CODES:
        error = 'A retriable HTTP error %d occurred:\n%s' % (e.resp.status,
                                                             e.content)
      else:
        raise
    except RETRIABLE_EXCEPTIONS, e:
      error = 'A retriable error occurred: %s' % e

    if error is not None:
      print error
      retry += 1
      if retry > MAX_RETRIES:
        exit('No longer attempting to retry.')

      max_sleep = 2 ** retry
      sleep_seconds = random.random() * max_sleep
      print 'Sleeping %f seconds and then retrying...' % sleep_seconds
      time.sleep(sleep_seconds)
  return returnvalue

if __name__ == '__main__':
  parser = argparse.ArgumentParser()
  parser.add_argument('--name', required=True, help='Name for the project')
  parser.add_argument('--category', default='20',
    help='Numeric video category. ' +
      'See https://developers.google.com/youtube/v3/docs/videoCategories/list')
  parser.add_argument('--keywords', help='Video keywords, comma separated', default='')
  parser.add_argument('--privacyStatus', choices=VALID_PRIVACY_STATUSES, default='unlisted', help='Video privacy status.')
  args = parser.parse_args()
  debug = args
  youtube = get_authenticated_service()

  try:
    initialize_upload(youtube, args)
  except HttpError, e:
    print 'An HTTP error %d occurred:\n%s' % (e.resp.status, e.content)